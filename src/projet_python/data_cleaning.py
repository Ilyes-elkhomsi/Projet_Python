import pandas as pd
from openpyxl import load_workbook, Workbook
from pathlib import Path
import requests
from io import BytesIO


def read_excel_with_openpyxl(path_or_url):
    if str(path_or_url).startswith("http"):
        response = requests.get(path_or_url)
        response.raise_for_status()
        file_content = BytesIO(response.content)
        wb = load_workbook(file_content, data_only=True)
    else:
        wb = load_workbook(path_or_url, data_only=True)

    ws = wb.active

    data = list(ws.values)
    headers = data[0]
    rows = data[1:]

    return pd.DataFrame(rows, columns=headers)

def write_df_with_openpyxl(df, path):
    Path(path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws.append(list(df.columns))

    for row in df.itertuples(index=False):
        ws.append(list(row))

    wb.save(path)


#DATA_URL = "https://minio.lab.sspcloud.fr/ilyeselkhomsii/Ligue_1.xlsx?X-Amz-Algorithm=AWS4-HMAC-SHA256&X-Amz-Content-Sha256=UNSIGNED-PAYLOAD&X-Amz-Credential=ZZTRF5W7NM05M4HW81QO%2F20260615%2Fus-east-1%2Fs3%2Faws4_request&X-Amz-Date=20260615T082005Z&X-Amz-Expires=86400&X-Amz-Security-Token=eyJhbGciOiJIUzUxMiIsInR5cCI6IkpXVCJ9.eyJhY2Nlc3NLZXkiOiJaWlRSRjVXN05NMDVNNEhXODFRTyIsImFsbG93ZWQtb3JpZ2lucyI6WyIqIl0sImF1ZCI6WyJtaW5pby1kYXRhbm9kZSIsIm9ueXhpYSIsImFjY291bnQiXSwiYXV0aF90aW1lIjoxNzgwOTAyNzk0LCJhenAiOiJvbnl4aWEiLCJlbWFpbCI6ImlseWVzLmVsLWtob21zaS5hdWRpdGV1ckBsZWNuYW0ubmV0IiwiZW1haWxfdmVyaWZpZWQiOnRydWUsImV4cCI6MTc4MjExMzc1OCwiZmFtaWx5X25hbWUiOiJFbCBraG9tc2kiLCJnaXZlbl9uYW1lIjoiSWx5ZXMiLCJncm91cHMiOlsiVVNFUl9PTllYSUEiXSwiaWF0IjoxNzgxNTA4OTU3LCJpc3MiOiJodHRwczovL2F1dGgubGFiLnNzcGNsb3VkLmZyL2F1dGgvcmVhbG1zL3NzcGNsb3VkIiwianRpIjoib25ydHJ0OmE2NjgwZjU3LTVhY2ItZDgzNi02M2Y0LTgyMWE4ODA3ZDNiYiIsImxvY2FsZSI6ImZyIiwibmFtZSI6IklseWVzIEVsIGtob21zaSIsInBvbGljeSI6InN0c29ubHkiLCJwcmVmZXJyZWRfdXNlcm5hbWUiOiJpbHllc2Vsa2hvbXNpaSIsInJlYWxtX2FjY2VzcyI6eyJyb2xlcyI6WyJvZmZsaW5lX2FjY2VzcyIsInVtYV9hdXRob3JpemF0aW9uIiwiZGVmYXVsdC1yb2xlcy1zc3BjbG91ZCJdfSwicmVzb3VyY2VfYWNjZXNzIjp7ImFjY291bnQiOnsicm9sZXMiOlsibWFuYWdlLWFjY291bnQiLCJtYW5hZ2UtYWNjb3VudC1saW5rcyIsInZpZXctcHJvZmlsZSJdfX0sInJvbGVzIjpbIm9mZmxpbmVfYWNjZXNzIiwidW1hX2F1dGhvcml6YXRpb24iLCJkZWZhdWx0LXJvbGVzLXNzcGNsb3VkIl0sInNjb3BlIjoib3BlbmlkIHByb2ZpbGUgZ3JvdXBzIGVtYWlsIiwic2lkIjoiMWRpSmltYjk4czY5ZDBMZVdwT1c5ZGR0Iiwic3ViIjoiOTQxNzM2ZjAtOTE5NS00OTcyLWI3NGUtOGEzN2JhODg0Yzg4IiwidHlwIjoiQmVhcmVyIn0.xoQkXq_e7mQBRq4fV4ZC2nzvmIINsaopnka-4gxZ2atSKSn-VNtiOd4ZFDlseoIIrrsB3CN344jGWLY9c75qiw&X-Amz-Signature=9193278d85039f4ba1d4db19c66592ce38a3d3a619b734512668883613046b58&X-Amz-SignedHeaders=host&x-amz-checksum-mode=ENABLED&x-id=GetObject"
DATA_URL = "https://minio.lab.sspcloud.fr/ilyeselkhomsii/Ligue_1/Ligue_1.xlsx"
df = read_excel_with_openpyxl(DATA_URL)

colonnes_numeriques = [
    "rang",
    "journees",
    "points",
    "buts_marques",
    "buts_encaisses",
    "clean_sheets",
    "xG",
    "xGA",
    "cartons_jaunes",
    "cartons_rouges",
    "buts_meilleur_buteur",
    "passes_decisives",
]

for col in colonnes_numeriques:
    if col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce")


def categorie_equipe(rang):
    if rang <= 5:
        return "Top 5"
    elif rang <= 12:
        return "Milieu"
    else:
        return "Bas de tableau"


df["efficacite_offensive"] = (df["buts_marques"] / df["xG"]).round(2)
df["solidite_defensive"] = (df["clean_sheets"] / df["journees"]).round(2)
df["indice_discipline"] = df["cartons_jaunes"] + 3 * df["cartons_rouges"]
df["categorie_equipe"] = df["rang"].apply(categorie_equipe)

write_df_with_openpyxl(df, "outputs/Ligue_1_clean.xlsx")

print("Nettoyage terminé : outputs/Ligue_1_clean.xlsx")