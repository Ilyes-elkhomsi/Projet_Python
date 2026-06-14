import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

df = pd.read_excel("outputs/Ligue_1_clean.xlsx")

def categorie_equipe(rang):
    if rang <= 5:
        return "Top 5"
    elif rang <= 12:
        return "Milieu"
    else:
        return "Bas de tableau"

df["categorie_equipe"] = df["rang"].apply(categorie_equipe)

output_file = "outputs/dashboard_ligue1.xlsx"

wb = Workbook()
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_data = wb.create_sheet("Data")


# DATA
for col_idx, col_name in enumerate(df.columns, start=1):
    ws_data.cell(row=1, column=col_idx, value=col_name)

for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        ws_data.cell(row=row_idx, column=col_idx, value=value)

col_map = {col: idx + 1 for idx, col in enumerate(df.columns)}

def data_cell(col_name, row):
    return f"Data!{get_column_letter(col_map[col_name])}{row}"


# STYLES
dark_fill = PatternFill("solid", fgColor="1F4E78")
blue_fill = PatternFill("solid", fgColor="D9EAF7")
kpi_fill = PatternFill("solid", fgColor="EAF3F8")
white_font = Font(color="FFFFFF", bold=True)
title_font = Font(size=20, bold=True, color="1F4E78")
subtitle_font = Font(size=11, italic=True, color="666666")
bold_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)


# TITRE
ws_dash.merge_cells("A1:L1")
ws_dash["A1"] = "Dashboard Ligue 1 - Analyse des facteurs de performance"
ws_dash["A1"].font = title_font
ws_dash["A1"].alignment = center

ws_dash.merge_cells("A2:L2")
ws_dash["A2"] = "Filtre interactif par catégorie d'équipe : Toutes, Top 5, Milieu ou Bas de tableau"
ws_dash["A2"].font = subtitle_font
ws_dash["A2"].alignment = center


# FILTRE
ws_dash["A4"] = "Catégorie sélectionnée"
ws_dash["A4"].font = bold_font
ws_dash["B4"] = "Toutes"

dv = DataValidation(
    type="list",
    formula1='"Toutes,Top 5,Milieu,Bas de tableau"',
    allow_blank=False
)
ws_dash.add_data_validation(dv)
dv.add(ws_dash["B4"])

ws_dash["B4"].fill = blue_fill
ws_dash["B4"].alignment = center
ws_dash["B4"].border = thin_border


# ZONE HELPER
helper_start_row = 90

headers = [
    "Equipe",
    "Points",
    "Buts marques",
    "Clean sheets",
    "xG",
    "Buts encaisses",
    "Meilleur buteur",
    "Buts meilleur buteur"
]

for col_idx, header in enumerate(headers, start=1):
    cell = ws_dash.cell(row=helper_start_row, column=col_idx, value=header)
    cell.font = bold_font
    cell.fill = blue_fill

for i in range(len(df)):
    excel_row = helper_start_row + 1 + i
    data_row = 2 + i

    condition = f'OR($B$4="Toutes",{data_cell("categorie_equipe", data_row)}=$B$4)'

    ws_dash.cell(row=excel_row, column=1, value=f'=IF({condition},{data_cell("equipe", data_row)},"")')
    ws_dash.cell(row=excel_row, column=2, value=f'=IF({condition},{data_cell("points", data_row)},"")')
    ws_dash.cell(row=excel_row, column=3, value=f'=IF({condition},{data_cell("buts_marques", data_row)},"")')
    ws_dash.cell(row=excel_row, column=4, value=f'=IF({condition},{data_cell("clean_sheets", data_row)},"")')
    ws_dash.cell(row=excel_row, column=5, value=f'=IF({condition},{data_cell("xG", data_row)},NA())')

    # Colonnes pour les KPI
    ws_dash.cell(row=excel_row, column=6, value=f'=IF({condition},{data_cell("buts_encaisses", data_row)},999)')
    ws_dash.cell(row=excel_row, column=7, value=f'=IF({condition},{data_cell("meilleur_buteur", data_row)},"")')
    ws_dash.cell(row=excel_row, column=8, value=f'=IF({condition},{data_cell("buts_meilleur_buteur", data_row)},0)')

# KPI
kpis = [
    (
        "Meilleure équipe",
        f'=IFERROR(INDEX($A${helper_start_row+1}:$A${helper_start_row+18},MATCH(MAX($B${helper_start_row+1}:$B${helper_start_row+18}),$B${helper_start_row+1}:$B${helper_start_row+18},0)),"-")'
    ),
    (
        "Meilleure attaque",
        f'=IFERROR(INDEX($A${helper_start_row+1}:$A${helper_start_row+18},MATCH(MAX($C${helper_start_row+1}:$C${helper_start_row+18}),$C${helper_start_row+1}:$C${helper_start_row+18},0)),"-")'
    ),
    (
        "Meilleure défense",
        f'=IFERROR(INDEX($A${helper_start_row+1}:$A${helper_start_row+18},MATCH(MIN($F${helper_start_row+1}:$F${helper_start_row+18}),$F${helper_start_row+1}:$F${helper_start_row+18},0)),"-")'
    ),
    (
        "Meilleur buteur",
        f'=IFERROR(INDEX($G${helper_start_row+1}:$G${helper_start_row+18},MATCH(MAX($H${helper_start_row+1}:$H${helper_start_row+18}),$H${helper_start_row+1}:$H${helper_start_row+18},0)),"-")'
    )
]

for i, (label, formula) in enumerate(kpis):
    col = 1 + i * 3
    ws_dash.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)
    ws_dash.cell(row=6, column=col, value=label)
    ws_dash.cell(row=6, column=col).fill = dark_fill
    ws_dash.cell(row=6, column=col).font = white_font
    ws_dash.cell(row=6, column=col).alignment = center

    ws_dash.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)
    ws_dash.cell(row=7, column=col, value=formula)
    ws_dash.cell(row=7, column=col).fill = kpi_fill
    ws_dash.cell(row=7, column=col).font = Font(size=12, bold=True)
    ws_dash.cell(row=7, column=col).alignment = center


# GRAPHIQUES
def graphique(title, data_col, position):
    chart = BarChart()

    chart.type = "bar"
    chart.style = 10
    chart.title = title

    # Axes corrects
    chart.x_axis.title = title
    chart.y_axis.title = "Équipes"

    chart.height = 8
    chart.width = 14
    chart.legend = None

    data = Reference(
        ws_dash,
        min_col=data_col,
        min_row=helper_start_row,
        max_row=helper_start_row + len(df)
    )

    cats = Reference(
        ws_dash,
        min_col=1,
        min_row=helper_start_row + 1,
        max_row=helper_start_row + len(df)
    )

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Étiquettes propres
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True       # affiche la valeur : 58, 56...
    chart.dataLabels.showCatName = False   # affiche l'équipe : Lyon, Lens...
    chart.dataLabels.showSerName = False  # évite "Points; Lyon; 58"

    ws_dash.add_chart(chart, position)

graphique("Points par équipe", 2, "A10")
graphique("Buts marqués par équipe", 3, "G10")
graphique("Clean sheets par équipe", 4, "A28")
graphique("xG par équipe", 5, "G28")

# ======================
# FEUILLE ANALYSE AVANCÉE
# ======================
ws_adv = wb.create_sheet("Analyse_avancee")

ws_adv.merge_cells("A1:L1")
ws_adv["A1"] = "Analyse avancée - Ligue 1"
ws_adv["A1"].font = title_font
ws_adv["A1"].alignment = center

ws_adv.merge_cells("A2:L2")
ws_adv["A2"] = "Analyse complémentaire : efficacité offensive, défense, discipline et domicile/extérieur"
ws_adv["A2"].font = subtitle_font
ws_adv["A2"].alignment = center

# Sécurité : création des variables si absentes
if "efficacite_offensive" not in df.columns:
    df["efficacite_offensive"] = (df["buts_marques"] / df["xG"]).round(2)

if "indice_discipline" not in df.columns:
    df["indice_discipline"] = df["cartons_jaunes"] + 3 * df["cartons_rouges"]

# Tables d'analyse
df_eff = df[["equipe", "efficacite_offensive"]].sort_values("efficacite_offensive", ascending=False)
df_def = df[["equipe", "buts_encaisses"]].sort_values("buts_encaisses", ascending=True)
df_disc = df[["equipe", "indice_discipline"]].sort_values("indice_discipline", ascending=False)
df_home = df[["equipe", "points_domicile", "points_exterieur"]].sort_values("points_domicile", ascending=False)

def write_table(ws, df_table, start_row, start_col):
    for j, col_name in enumerate(df_table.columns, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=col_name)
        cell.fill = blue_fill
        cell.font = bold_font

    for i, row in enumerate(df_table.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row, start=start_col):
            ws.cell(row=i, column=j, value=value)

def add_simple_chart(ws, title, start_row, start_col, value_col, position):
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.x_axis.title = title
    chart.y_axis.title = "Équipes"
    chart.height = 8
    chart.width = 13
    chart.legend = None

    data = Reference(ws, min_col=value_col, min_row=start_row, max_row=start_row + 18)
    cats = Reference(ws, min_col=start_col, min_row=start_row + 1, max_row=start_row + 18)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False

    ws.add_chart(chart, position)

# Écriture des tables
write_table(ws_adv, df_eff, 60, 1)
write_table(ws_adv, df_def, 60,5)
write_table(ws_adv, df_disc, 85,1)
write_table(ws_adv, df_home,85,5)

# Graphiques analyse avancée
add_simple_chart(ws_adv, "Efficacité offensive", 60,1, 2, "A8")
add_simple_chart(ws_adv, "Buts encaissés",60, 5, 6, "G8")
add_simple_chart(ws_adv, "Indice discipline", 85,1,2,"A28")

# Graphique domicile / extérieur
chart_home = BarChart()
chart_home.type = "bar"
chart_home.style = 10
chart_home.title = "Points domicile vs extérieur"
chart_home.x_axis.title = "Points"
chart_home.y_axis.title = "Équipes"
chart_home.height = 8
chart_home.width = 13

data = Reference(ws_adv, min_col=6, max_col=7, min_row=85, max_row=103)
cats = Reference(ws_adv, min_col=5, min_row=86, max_row=103)

ws_adv.add_chart(chart_home, "G25")

chart_home.add_data(data, titles_from_data=True)
chart_home.set_categories(cats)


for col in range(1, 13):
    ws_adv.column_dimensions[get_column_letter(col)].width = 18

# MISE EN FORME
for col in range(1, 13):
    ws_dash.column_dimensions[get_column_letter(col)].width = 15

for col in range(1, len(df.columns) + 1):
    ws_data.column_dimensions[get_column_letter(col)].width = 18

# On masque la feuille Data mais on garde la zone helper visible pour l’instant
ws_data.sheet_state = "hidden"

wb.save(output_file)

print("Dashboard généré automatiquement :", output_file)